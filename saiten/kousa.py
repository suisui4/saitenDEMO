import tkinter
import tkinter as tk
from tkinter import messagebox
from tkinter.constants import TRUE
from PIL import Image, ImageTk, ImageDraw, ImageFont  # 外部ライブラリ
import csv
import glob
import os
import sys
import shutil
import pathlib
import imghdr
import openpyxl
import pandas as pd
import numpy as np
import cv2 
from keras.models import load_model
from keras.models import model_from_json


global canvas1
global img_resized
global img_tk
global Giri_cutter
global root
global topimg
global topfig

global RESIZE_RETIO  # 縮小倍率の規定

os.environ['TF_CPP_MIN_LOG_LEVEL'] = '2'
window_h = 700
window_w = int(window_h * 1.7)
fig_area_w = int(window_h * 1)

global qCnt
qCnt = 0

# 画像パスの取得
# https://msteacher.hatenablog.jp/entry/2020/06/27/170529
def resource_path(relative_path):
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.join(os.path.abspath("."), relative_path)



# 初回起動時にフォルダを展開する。
def initDir():
    os.makedirs("./setting/input", exist_ok=True)
    os.makedirs("./setting/output", exist_ok=True)
    f = open('setting/ini.csv', 'w')  # 既存でないファイル名を作成してください
    writer = csv.writer(f, lineterminator='\n')  # 行末は改行
    writer.writerow(["tag", "start_x", "start_y", "end_x", "end_y"])
    f.close()

# 与えられたフォルダの全てのファイル(フルパス)をソートしてリストで返す
# 拡張子が画像かどうかも判別し、画像のパスのみを返す。


def get_sorted_files(dir_path):
    all_sorted = sorted(glob.glob(dir_path))
    fig_sorted = [s for s in all_sorted if s.endswith(
        ('jpg', "jpeg", "png", "PNG", "JPEG", "JPG", "gif",'bmp'))]
    return fig_sorted


# ドラッグ開始した時のイベント - - - - - - - - - - - - - - - - - - - - - - - - - -
def start_point_get(event):
    global start_x, start_y  # グローバル変数に書き込みを行なうため宣言

    canvas1.delete("rectTmp")  # すでに"rectTmp"タグの図形があれば削除

    # canvas1上に四角形を描画（rectangleは矩形の意味）
    canvas1.create_rectangle(event.x,
                             event.y,
                             event.x + 1,
                             event.y + 1,
                             outline="red",
                             tag="rectTmp")
    # グローバル変数に座標を格納
    start_x, start_y = event.x, event.y

# ドラッグ中のイベント - - - - - - - - - - - - - - - - - - - - - - - - - -


def rect_drawing(event):

    # ドラッグ中のマウスポインタが領域外に出た時の処理
    if event.x < 0:
        end_x = 0
    else:
        end_x = min(img_resized.width, event.x)
    if event.y < 0:
        end_y = 0
    else:
        end_y = min(img_resized.height, event.y)

    # "rectTmp"タグの画像を再描画
    canvas1.coords("rectTmp", start_x, start_y, end_x, end_y)

# ドラッグを離したときのイベント - - - - - - - - - - - - - - - - - - - - - - - - - -


def release_action(event):
    global qCnt

    if qCnt == 0:
        pos = canvas1.bbox("rectTmp")

        # canvas1上に四角形を描画（rectangleは矩形の意味）
        create_rectangle_alpha(pos[0], pos[1], pos[2], pos[3],
                               fill="green",
                               alpha=0.3,
                               tag="nameBox"
                               )

        canvas1.create_text(
            (pos[0] + pos[2]) / 2, (pos[1] + pos[3]) / 2,
            text="name",
            tag="nameText"
        )

        # "rectTmp"タグの画像の座標を元の縮尺に戻して取得
        start_x, start_y, end_x, end_y = [
            round(n * RESIZE_RETIO) for n in canvas1.coords("rectTmp")
        ]
        with open('setting/ini.csv', 'a') as f:
            writer = csv.writer(f, lineterminator='\n')  # 行末は改行
            writer.writerow(["name", start_x, start_y, end_x, end_y])

    else:
        pos = canvas1.bbox("rectTmp")
        # canvas1上に四角形を描画（rectangleは矩形の意味）
        create_rectangle_alpha(pos[0], pos[1], pos[2], pos[3],
                               fill="red",
                               alpha=0.3,
                               tag="qBox" + str(qCnt)
                               )
        canvas1.create_text(
            (pos[0] + pos[2]) / 2, (pos[1] + pos[3]) / 2,
            text="Q_" + str(qCnt),
            tag="qText" + str(qCnt)
        )

        # "rectTmp"タグの画像の座標を元の縮尺に戻して取得
        start_x, start_y, end_x, end_y = [
            round(n * RESIZE_RETIO) for n in canvas1.coords("rectTmp")
        ]
        with open('setting/ini.csv', 'a') as f:
            writer = csv.writer(f, lineterminator='\n')  # 行末は改行
            writer.writerow(["Q_" + str(qCnt).zfill(4),
                            start_x, start_y, end_x, end_y])

    qCnt = qCnt + 1


# 透過画像の作成 - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
# https://stackoverflow.com/questions/54637795/how-to-make-a-tkinter-canvas-rectangle-transparent/54645103
# 透過画像を削除するときは、imagesの配列から消す。
images = []  # to hold the newly created image


def create_rectangle_alpha(x1, y1, x2, y2, **kwargs):
    if 'alpha' in kwargs:
        alpha = int(kwargs.pop('alpha') * 255)
        fill = kwargs.pop('fill')
        fill = Giri_cutter.winfo_rgb(fill) + (alpha,)
        image = Image.new('RGBA', (x2-x1, y2-y1), fill)
        images.append(ImageTk.PhotoImage(image, master=Giri_cutter))
        canvas1.create_image(x1, y1, image=images[-1], anchor='nw')
    canvas1.create_rectangle(x1, y1, x2, y2, **kwargs)


def back_one():
    global qCnt
    if qCnt == 0:
        return
    qCnt = qCnt - 1
    # タグに基づいて画像を削除
    if qCnt == 0:
        canvas1.delete("nameBox", "nameText", "rectTmp")
        images.pop(-1)
    else:
        canvas1.delete("qBox" + str(qCnt), "qText" + str(qCnt), "rectTmp")
        images.pop(-1)
    # csvの最終行を削除
    readFile = open("setting/ini.csv")
    lines = readFile.readlines()
    readFile.close()
    w = open("setting/ini.csv", 'w')
    w.writelines([item for item in lines[:-1]])
    w.close()


def trim_fin():
    global Giri_cutter
    ret = messagebox.askyesno('終了します', '斬り方を決定し、ホームに戻っても良いですか？')
    if ret == True:
        cur = os.getcwd()
        beforePath = cur + "/setting/ini.csv"
        afterPath = cur + "/setting/trimData.csv"
        shutil.move(beforePath, afterPath)
        Giri_cutter.destroy()


def setting_ck():
    if not os.path.exists("./setting/"):
        ret = messagebox.askyesno(
            '初回起動です', '採点のために、いくつかのフォルダーをこのファイルと同じ場所に作成します。\nよろしいですか？')
        if ret == True:
            initDir()
            messagebox.showinfo(
                '準備ができました。', '解答用紙を、setting/input の中に保存してください。jpeg または png に対応しています。')
        else:
            # メッセージボックス（情報）
            messagebox.showinfo('終了', 'フォルダは作成しません。')
    else:
        messagebox.showinfo(
            '確認', '初期設定は完了しています。解答用紙を、setting/inputに入れてから、解答用紙分割をしてください。')


def input_ck():
    # 表示する画像の取得
    files = get_sorted_files(os.getcwd() + "/setting/input/*")
    if not files:
        # メッセージボックス（警告）
        messagebox.showerror(
            "エラー", "setting/inputの中に、解答用紙のデータが存在しません。画像を入れてから、また開いてね。")
    else:
        GiriActivate()

def release_action2(event):
    global qCnt

    pos = canvas1.bbox("rectTmp")
        # canvas1上に四角形を描画（rectangleは矩形の意味）
    create_rectangle_alpha(pos[0], pos[1], pos[2], pos[3],
                               fill="red",
                               alpha=0.3,
                               tag="qBox" + str(qCnt)
                               )
    canvas1.create_text(
            (pos[0] + pos[2]) / 2, (pos[1] + pos[3]) / 2,
            text="kanten",
            tag="qText" + str(qCnt)
        )

        # "rectTmp"タグの画像の座標を元の縮尺に戻して取得
    start_x, start_y, end_x, end_y = [
            round(n * RESIZE_RETIO) for n in canvas1.coords("rectTmp")
        ]
    with open('setting/ini2.csv', 'a') as f:
        writer = csv.writer(f, lineterminator='\n')  # 行末は改行
        writer.writerow(["kanten",
                            start_x, start_y, end_x, end_y])

    qCnt = qCnt + 1
        
# def release_action2(event):
#     global qCnt
#     pos = canvas1.bbox("rectTmp")
#     # canvas1上に四角形を描画（rectangleは矩形の意味）
#     create_rectangle_alpha(pos[0], pos[1], pos[2], pos[3],
#                            fill="red",
#                            alpha=0.3,
#                            tag="qBox" + str(qCnt)
#                            )
#     canvas1.create_text(
#         (pos[0] + pos[2]) / 2, (pos[1] + pos[3]) / 2,
#             text="kanten",
#         tag="qText" + str(qCnt))

#     # "rectTmp"タグの画像の座標を元の縮尺に戻して取得
#     start_x, start_y, end_x, end_y = [
#         round(n * RESIZE_RETIO) for n in canvas1.coords("rectTmp")
#     ]
#     with open('setting/kanten.csv', 'a') as f:
#         writer = csv.writer(f, lineterminator='\n')  # 行末は改行
#         writer.writerow(["kanten",
#                         start_x, start_y, end_x, end_y])

#     qCnt = qCnt + 1  

def trim_fin2():
    global Giri_cutter
    ret = messagebox.askyesno('終了します', '観点別OKですか？')
    if ret == True:
        cur = os.getcwd()
        beforePath = cur + "/setting/ini2.csv"
        afterPath = cur + "/setting/kanten_matrix.csv"
        shutil.move(beforePath, afterPath)
        Giri_cutter.destroy()    

def back_one2():
    global qCnt
    if qCnt == 0:
        return
    qCnt = qCnt - 1
    # タグに基づいて画像を削除
    if qCnt == 0:
        canvas1.delete("nameBox", "nameText", "rectTmp")
        images.pop(-1)
    else:
        canvas1.delete("qBox" + str(qCnt), "qText" + str(qCnt), "rectTmp")
        images.pop(-1)
    # csvの最終行を削除
    readFile = open("setting/kanten_matrix.csv")
    lines = readFile.readlines()
    readFile.close()
    w = open("setting/kanten_matrix.csv", 'w')
    w.writelines([item for item in lines[:-1]])
    w.close()
    
def Kantenbetu():
    global RESIZE_RETIO
    global img_resized
    global canvas1
    global img_tk
    global Giri_cutter
    global qCnt

    def toTop():
        global qCnt
        ret = messagebox.askyesno(
            '保存しません', '作業中のデータは保存されません。\n画面を移動しますか？')
        if ret == True:
            qCnt = 0
            Giri_cutter.destroy()
        else:
            pass

    # 表示する画像の取得
    files = get_sorted_files(os.getcwd() + "/setting/input/*")
    name_lists=[]
    for file in files:
        name = file.split("t\\")[1]
        name_lists.append(name)
    img = Image.open(files[0])
    # 画面サイズに合わせて画像をリサイズする
    # 画像サイズが縦か横かに合わせて、RESIZE_RETIOを決める。
    w, h = img.size
    if w >= h:
        if w <= fig_area_w:
            RESIZE_RETIO = 1
        else:
            RESIZE_RETIO = w / fig_area_w
    else:
        if h <= window_h:
            RESIZE_RETIO = 1
        else:
            RESIZE_RETIO = h / window_h

    # 画像リサイズ
    img_resized = img.resize(size=(int(img.width / RESIZE_RETIO),
                                   int(img.height / RESIZE_RETIO)),
                             resample=Image.BILINEAR)
    Giri_cutter = tkinter.Tk()
    Giri_cutter.geometry(str(window_w) + "x" + str(window_h))
    Giri_cutter.title("観点別データ作成")
    cutting_frame = tkinter.Frame(Giri_cutter)
    cutting_frame.pack()
    canvas_frame = tkinter.Frame(cutting_frame)
    canvas_frame.grid(column=0, row=0)
    button_frame = tkinter.Frame(cutting_frame)
    button_frame.grid(column=1, row=0)
    # tkinterで表示できるように画像変換
    img_tk = ImageTk.PhotoImage(img_resized, master=Giri_cutter)
    # Canvasウィジェットの描画
    canvas1 = tkinter.Canvas(canvas_frame,
                             bg="black",
                             width=img_resized.width,
                             height=img_resized.height,
                             highlightthickness=0)
    # Canvasウィジェットに取得した画像を描画
    canvas1.create_image(0, 0, image=img_tk, anchor=tkinter.NW)

    # Canvasウィジェットを配置し、各種イベントを設定
    canvas1.pack()

    # 戻るボタン
    backB = tkinter.Button(
        button_frame, text='一つ前に戻る', command=back_one2, width=20, height=4).pack()

    # 入力完了
    finB = tkinter.Button(
        button_frame, text='入力完了\n(保存して戻る)', command=trim_fin2, width=20, height=4).pack()
    topB = tkinter.Button(
        button_frame, text='topに戻る\n(必ず選択がない状態で終わるように)', command=toTop, width=20, height=4).pack()

    canvas1.bind("<ButtonPress-1>", start_point_get)
    canvas1.bind("<Button1-Motion>", rect_drawing)
    canvas1.bind("<ButtonRelease-1>", release_action2)
    Giri_cutter.mainloop()              
        
        
        
def GiriActivate():
    global RESIZE_RETIO
    global img_resized
    global canvas1
    global img_tk
    global Giri_cutter
    global qCnt

    def toTop():
        global qCnt
        ret = messagebox.askyesno(
            '保存しません', '作業中のデータは保存されません。\n画面を移動しますか？')
        if ret == True:
            qCnt = 0
            Giri_cutter.destroy()
        else:
            pass

    # 表示する画像の取得
    files = get_sorted_files(os.getcwd() + "/setting/input/*")
    name_lists=[]
    for file in files:
        name = file.split("t/")[1]
        name_lists.append(name)
    print(name_lists)
    # ini.csvは、起動のたびに初期化する。
    f = open('setting/ini.csv', 'w')  # 既存でないファイル名を作成してください
    writer = csv.writer(f, lineterminator='\n')  # 行末は改行
    writer.writerow(["tag", "start_x", "start_y", "end_x", "end_y"])
    f.close()

    img = Image.open(files[0])

    # 画面サイズに合わせて画像をリサイズする
    # 画像サイズが縦か横かに合わせて、RESIZE_RETIOを決める。
    w, h = img.size
    if w >= h:
        if w <= fig_area_w:
            RESIZE_RETIO = 1
        else:
            RESIZE_RETIO = w / fig_area_w
    else:
        if h <= window_h:
            RESIZE_RETIO = 1
        else:
            RESIZE_RETIO = h / window_h

    # 画像リサイズ
    img_resized = img.resize(size=(int(img.width / RESIZE_RETIO),
                                   int(img.height / RESIZE_RETIO)),
                             resample=Image.BILINEAR)

    Giri_cutter = tkinter.Tk()
    Giri_cutter.geometry(str(window_w) + "x" + str(window_h))
    Giri_cutter.title("名前を選択")

    cutting_frame = tkinter.Frame(Giri_cutter)
    cutting_frame.pack()
    canvas_frame = tkinter.Frame(cutting_frame)
    canvas_frame.grid(column=0, row=0)
    button_frame = tkinter.Frame(cutting_frame)
    button_frame.grid(column=1, row=0)

    # tkinterで表示できるように画像変換
    img_tk = ImageTk.PhotoImage(img_resized, master=Giri_cutter)

    # Canvasウィジェットの描画
    canvas1 = tkinter.Canvas(canvas_frame,
                             bg="black",
                             width=img_resized.width,
                             height=img_resized.height,
                             highlightthickness=0)
    # Canvasウィジェットに取得した画像を描画
    canvas1.create_image(0, 0, image=img_tk, anchor=tkinter.NW)

    # Canvasウィジェットを配置し、各種イベントを設定
    canvas1.pack()

    # 戻るボタン
    backB = tkinter.Button(
        button_frame, text='一つ前に戻る', command=back_one, width=20, height=4).pack()

    # 入力完了
    finB = tkinter.Button(
        button_frame, text='入力完了\n(保存して戻る)', command=trim_fin, width=20, height=4).pack()
    topB = tkinter.Button(
        button_frame, text='topに戻る\n(保存はされません)', command=toTop, width=20, height=4).pack()

    canvas1.bind("<ButtonPress-1>", start_point_get)
    canvas1.bind("<Button1-Motion>", rect_drawing)
    canvas1.bind("<ButtonRelease-1>", release_action)
    Giri_cutter.mainloop()


def trimck():
    ret = messagebox.askyesno(
        '全員の解答用紙を斬ります。\n 受験者が100人以上いると5分ほど時間がかかります。\n現在のoutputは全て消えます。')
    if ret == True:
        allTrim()
    else:
        pass


def allTrim():
    # トリミング前の画像の格納先
    ORIGINAL_FILE_DIR = "./setting/input"
    # トリミング後の画像の格納先
    TRIMMED_FILE_DIR = "./setting/output"

    def readCSV():
        # もしcsvが無ければ、全部止める
        if os.path.isfile("./setting/trimData.csv") == False:
            return 0
        else:
            with open('./setting/trimData.csv') as f:
                reader = csv.reader(f)
                data = [row for row in reader]
                data.pop(0)
                return data

    data = readCSV()

    try:
        shutil.rmtree("./setting/output")
    except OSError as err:
        pass

    if data == 0:
        messagebox.showinfo('終了', 'どうやって斬ればいいかわかりません。\nまずはどこを斬るかを決めてください。')
        return 0






    # 画像ファイル名を取得
    files = os.listdir(ORIGINAL_FILE_DIR)
    # 特定の拡張子のファイルだけを採用。実際に加工するファイルの拡張子に合わせる
    files = [name for name in files if name.split(
        ".")[-1] in ['jpg', "jpeg", "png", "PNG", "JPEG", "JPG", "gif","bmp"]]

    try:
        for val in files:
            # オリジナル画像へのパス
            path = ORIGINAL_FILE_DIR + "/" + val
            # トリミングされたimageオブジェクトを取得
            im = Image.open(path)
            print(val + "を斬ります" )
            for pos in data:
                # 出力フォルダのパス
                title , left , top , right , bottom = pos
                outputDir = TRIMMED_FILE_DIR + "/" + title
                # もしトリミング後の画像の格納先が存在しなければ作る
                if os.path.isdir(outputDir) == False:
                    os.makedirs(outputDir)


                im_trimmed = im.crop((int(left), int(top), int(right), int(bottom)))
                # qualityは95より大きい値は推奨されていないらしい
                im_trimmed.save(outputDir + "/" + val, quality=95)

                print("___"+ title + "を斬り取りました。" )
            print("********************************")
    except:
        messagebox.showinfo(
            'エラー', 'エラーが検出されました。中断します。\n\n' + str(sys.stderr))
        try:
            shutil.rmtree("./setting/output")
        except OSError as err:
            pass
        return 0


    # nameフォルダの中身をリサイズ
    # maxheight以上のときは、小さくする。
    maxheight = 50
    files = glob.glob("./setting/output/name/*")
    img = Image.open(files[0])
    namew, nameh = img.size
    if nameh > maxheight:
        rr = nameh / maxheight
        for f in files:
            img = Image.open(f)
            img = img.resize((int(namew / rr), int(nameh/rr)))
            img.save(f)

    output_name_sh()
    # messagebox.showinfo('斬りました', '全員分の解答用紙を斬りました。')


def exitGiri():
    sys.exit()


def info():
    messagebox.showinfo(
        "はじめに", "オンラインヘルプをご覧ください。\n https://phys-ken.github.io/saitenGiri2021/")


def outputXlsx():
    try:
        saiten2xlsx()
        messagebox.showinfo("確認", "setting/saiten.xlsxに、採点結果を書き込みました。")
    except:
        messagebox.showerror("エラー", "うまくいきませんでした...")


def saiten2xlsx():
    def readCSV():
        # もしcsvが無ければ、全部止める
        if os.path.isfile("./setting/trimData.csv") == False:
            return 0
        else:
            with open('./setting/trimData.csv') as f:
                reader = csv.reader(f)
                data = [row for row in reader]
                data.pop(0)
                return data

    def setTensu(figname, qname, tensu):

        qCol = int(qname[-3:]) + 3
        ws.cell(1, qCol + 1).value = qname

        MIN_COL = 1
        MIN_ROW = 2

        MAX_COL = 1
        MAX_ROW = ws.max_row

        # 範囲データを順次処理
        for row in ws.iter_rows(min_col=MIN_COL, min_row=MIN_ROW, max_col=MAX_COL, max_row=MAX_ROW):
            for cell in row:
                try:
                    # 該当セルの値取得
                    cell_value = cell.value
                    if figname == cell_value:
                        o = cell.offset(0, qCol)
                        try:
                            o.value = int(tensu)
                        except:
                            o.value = tensu
                except:
                    pass

    data = readCSV()

    xlPath = "./setting/saiten.xlsx"
    wb = openpyxl.load_workbook(xlPath)
    ws = wb["採点シート"]

    while data:
        title, left, top, right, bottom = data.pop(0)
        if title == "name":
            continue
        qpath = "./setting/output/" + title
        for curDir, dirs, files in os.walk(qpath):
            if files:
                for f in files:
                    tensu = os.path.basename(os.path.dirname(curDir + "/" + f))
                    if not tensu == title:
                        setTensu(figname=f, qname=title, tensu=tensu)
                    else:
                        setTensu(figname=f, qname=title, tensu="未")
    wb.save(xlPath)


def saitenSelect():
    def show_selection():
        for i in lb.curselection():
            print(lb.get(i))
            siwakeApp(str(lb.get(i)))
            selectQ.destroy()
    
    def show_all_saiten():
        global folda
        for i in lb.curselection():
            folda = str(lb.get(i))
            pd.options.display.precision = 0
            df_ans_summmary_fix=pd.read_csv("./seitoPDF_CSV/df_ans_summmary_fix.csv",index_col = 0)
            if df_ans_summmary_fix[folda].isnull().all() == True:
                df_ans_summmary = pd.read_csv('./seitoPDF_CSV/df_ans_summmary.csv',index_col = 0)
                list_ = df_ans_summmary[[folda]].dropna()[folda].to_list()
                answer_text_list=[round(i) if type(i) is float else i for i in list_ ]
            else:
                list_ = df_ans_summmary_fix[[folda]].dropna()[folda].to_list()
                answer_text_list=[round(i) if type(i) is float else i for i in list_ ]
            summary_saiten(folda,answer_text_list)
#             selectQ.destroy()

    def backTop():
        selectQ.destroy()

    selectQ = tkinter.Tk()
    selectQ.geometry("500x500")
    selectQ.title("採点する問題を選ぶ")
    # macにおける、.DS_storeを無視してカウントする。
    maxNinzu = len([f for f in next(os.walk("./setting/input/"))[2] if not f.startswith('.')])

    # outputの中のフォルダを取得
    path = "./setting/output/"
    files = os.listdir(path)
    files_dir = [f for f in files if os.path.isdir(os.path.join(path, f))]
    files_dir.sort()
    lb = tkinter.Listbox(selectQ, selectmode='single', height=20, width=20)
    clcounter = 0
    for i in files_dir:
        if not i == "name":
            misaiten = len([f for f in next(os.walk("./setting/output/" + i))[2] if not f.startswith('.')])
            lb.insert(tkinter.END, i)
            if misaiten == maxNinzu:
                lb.itemconfig(clcounter, {'bg': 'white'})
            elif misaiten == 0 or misaiten == 1:
                lb.itemconfig(clcounter,  {'bg': 'gray'})
            else:
                lb.itemconfig(clcounter,  {'bg': 'pale green'})
            clcounter = clcounter + 1

    lb.grid(row=0, column=0)
    # Scrollbar
    scrollbar = tkinter.Scrollbar(
        selectQ,
        orient=tkinter.VERTICAL,
        command=lb.yview)
    lb['yscrollcommand'] = scrollbar.set
    scrollbar.grid(row=0, column=1,  sticky=(tkinter.N, tkinter.S, tkinter.W))

    button_frame = tkinter.Frame(selectQ)
    button_frame.grid(row=0, column=1, sticky=tkinter.W +
                      tkinter.E + tkinter.N + tkinter.S, padx=30, pady=30)

    siroKaisetsu = tkinter.Label(button_frame, text="未採点", bg="white").pack(
        side=tkinter.TOP, fill=tkinter.X)
    midoriKaisetsu = tkinter.Label(button_frame, text="採点中", bg="pale green").pack(
        side=tkinter.TOP, fill=tkinter.X)
    grayKaisetsu = tkinter.Label(button_frame, text="採点終了", bg="gray").pack(
        side=tkinter.TOP, fill=tkinter.X)

    button1 = tkinter.Button(
        button_frame, text='採点する', width=15, height=3,
        command=lambda: show_selection()).pack(expand=True)
    
    button2 = tkinter.Button(
        button_frame, text='全部一括採点する', width=15, height=3,
        command=lambda: show_all_saiten()).pack(expand=True)

    totopB = tkinter.Button(
        button_frame, text='Topに戻る', width=15, height=3,
        command=backTop).pack()

    selectQ.mainloop


def folder_walker(folder_path, recursive=False, file_ext=".*"):
    """
    指定されたフォルダのファイル一覧を取得する。
    引数を指定することで再帰的にも、非再帰的にも取得可能。

    Parameters
    ----------
    folder_path : str
        対象のフォルダパス
    recursive : bool
        再帰的に取得するか否か。既定値はTrueで再帰的に取得する。
    file_ext : str
        読み込むファイルの拡張子を指定。例：".jpg"のようにピリオドが必要。既定値は".*"で指定なし
    """

    p = pathlib.Path(folder_path)

    if recursive:
        return list(p.glob("**/*" + file_ext))  # **/*で再帰的にファイルを取得
    else:
        return list(p.glob("*" + file_ext))  # 再帰的にファイル取得しない


# ファイル読み込み - - - - - - - - - - - - - - - - - - - - - - - -
def load_file(Qnum):

    global img_num, item, dir_name

    # ファイルを読み込み
    tex_var.set("ファイルを読み込んでいます...")
    dir_name = "./setting/output/" + Qnum
    if not dir_name == None:
        file_list = folder_walker(dir_name)

    # ファイルから読み込める画像をリストに列挙
    for f in file_list:
        try:
            print("__読み込み中..." + str(f))
            img_lst.append(Image.open(f))
            filename_lst.append(f)
        except:
            pass

    if not img_lst:
        tex_var.set("読み込む画像がありません。\n採点は終了しています。")

    # ウィンドウサイズに合わせてキャンバスサイズを再定義
    # window_resize()

    # 画像変換
    for f in img_lst:

        # キャンバス内に収まるようリサイズ
        resized_img = img_resize_for_canvas(f, image_canvas, expand=True)

        # tkinterで表示できるように画像変換
        tk_img_lst.append(ImageTk.PhotoImage(
            image=resized_img, master=image_canvas))

    # キャンバスの中心を取得
    c_width_half = round(int(image_canvas["width"]) / 2)
    c_height_half = round(int(image_canvas["height"]) / 2)

    # キャンバスに表示
    img_num = 0
    item = image_canvas.create_image(
        c_width_half, c_height_half,  image=tk_img_lst[0], anchor=tkinter.CENTER)
    # ラベルの書き換え
    tex_var.set(filename_lst[img_num])
    saitenCount.set(str(img_num+1) + "/" + str(len(filename_lst)))

    # 仕分け実行ボタンの配置
    assort_btn.pack(expand=True)

# 次の画像へ - - - - - - - - - - - - - - - - - - - - - - - -


def next_img(event):
    global img_num

    # 読み込んでいる画像の数を取得
    img_count = len(tk_img_lst)

    # 画像が最後でないか判定
    if img_num >= img_count - 1:
        pass
    else:
        # 表示中の画像No.を更新して表示
        img_num += 1
        image_canvas.itemconfig(item, image=tk_img_lst[img_num])
        # ラベルの書き換え
        tex_var.set(filename_lst[img_num])
        saitenCount.set(str(img_num+1) + "/" + str(len(filename_lst)))
        # ラベリングを表示
        if filename_lst[img_num] in assort_dict:
            assort_t_var.set(assort_dict[filename_lst[img_num]])
        else:
            assort_t_var.set("")


# 前の画像へ - - - - - - - - - - - - - - - - - - - - - - - -
def prev_img(event):
    global img_num

    # 画像が最初でないか判定
    if img_num <= 0:
        pass
    else:
        # 表示中の画像No.を更新して表示
        img_num -= 1
        image_canvas.itemconfig(item, image=tk_img_lst[img_num])
        # ラベルの書き換え
        tex_var.set(filename_lst[img_num])
        saitenCount.set(str(img_num+1) + "/" + str(len(filename_lst)))
        # ラベリングを表示
        if filename_lst[img_num] in assort_dict:
            assort_t_var.set(assort_dict[filename_lst[img_num]])
        else:
            assort_t_var.set("")


# ウィンドウサイズからキャンバスサイズを再定義 - - - - - - - - - - - - - - - - -
def window_resize():

    image_canvas["width"] = image_canvas.winfo_width()
    image_canvas["height"] = image_canvas.winfo_height()


# キャンバスサイズに合わせて画像を縮小 - - - - - - - - - - - - - - - - - - - -
def img_resize_for_canvas(img, canvas, expand=False):

    size_retio_w = int(canvas["width"]) / img.width
    size_retio_h = int(canvas["height"]) / img.height

    if expand == True:
        size_retio = min(size_retio_w, size_retio_h)
    else:
        size_retio = min(size_retio_w, size_retio_h, 1)

    resized_img = img.resize((round(img.width * size_retio),
                              round(img.height * size_retio)))
    return resized_img

# 画像表示 - - - - - - - - - - - - - - - - - - - - - - - -


def image_show(event):
    img_lst[img_num].show()


# 画像に対しラベリング - - - - - - - - - - - - - - - - - - - - - - - -
def file_assort(event):
    tokutenList = []
    if cbln0.get():
        tokutenList.append("0")
    if cbln1.get():
        tokutenList.append("1")
    if cbln2.get():
        tokutenList.append("2")
    if cbln3.get():
        tokutenList.append("3")
    if cbln4.get():
        tokutenList.append("4")
    if cbln5.get():
        tokutenList.append("5")
    if cbln6.get():
        tokutenList.append("6")
    if cbln7.get():
        tokutenList.append("7")
    if cbln8.get():
        tokutenList.append("8")
    if cbln9.get():
        tokutenList.append("9")
    print("入力可能な点数は" + str(tokutenList))

    if str(event.keysym) in tokutenList:
        assort_dict[filename_lst[img_num]] = str(event.keysym)
    elif str(event.keysym) == "space":
        assort_dict[filename_lst[img_num]] = str("skip")
    elif str(event.keysym) in  ["0", "1", "2", "3", "4", "5", "6", "7", "8", "9"]:
        assort_dict[filename_lst[img_num]] = str("その点数は入力できません。\n右のチェックを確認してください。")
    else:
        assort_dict[filename_lst[img_num]] = str("そのキーは対応してません。")

    # ラベリングを表示
    if filename_lst[img_num] in assort_dict:
        assort_t_var.set(assort_dict[filename_lst[img_num]])
    else:
        assort_t_var.set("")

    print(assort_dict[filename_lst[img_num]])


# フォルダ分け実行 - - - - - - - - - - - - - - - - - - - - - - - -
def assort_go(event):

    global f_dir

    for f in assort_dict:
        # 仕分け前後のファイル名・フォルダ名を取得
        # assort_dict[f]が[0~1]なら、フォルダを作る。

        print(assort_dict[f])
        if assort_dict[f] in ["0", "1", "2", "3", "4", "5", "6", "7", "8", "9"]:
            f_dir = os.path.dirname(f)
            f_basename = os.path.basename(f)
            new_dir = os.path.join(f_dir, assort_dict[f])
            new_path = os.path.join(new_dir, f_basename)

            # ディレクトリの存在チェック
            if not os.path.exists(new_dir):
                os.mkdir(new_dir)
            # ファイルの移動実行
            shutil.move(f, new_path)

            print(new_path)
        else:
            pass
    saiten2xlsx()
    messagebox.showinfo("採点保存", "ここまでの採点結果を保存しました。\nskipした項目は、採点されていません。")
    siwake_win.destroy()


def siwakeApp(Qnum):
    def exit_siwake():
        ret = messagebox.askyesno('終了します', '採点を中断し、ホームに戻っても良いですか？')
        if ret == True:
            siwake_win.destroy()

    # グローバル変数
    global img_lst, tk_img_lst
    global filename_lst
    global assort_file_list
    global assort_dict

    global tex_var
    global image_canvas
    global assort_btn
    global assort_t_var

    global img_num
    global f_basename
    global siwake_win
    global saitenCount

    global cbln0
    global cbln1
    global cbln2
    global cbln3
    global cbln4
    global cbln5
    global cbln6
    global cbln7
    global cbln8
    global cbln9

    img_lst, tk_img_lst = [], []
    filename_lst = []
    assort_file_list = []
    assort_dict = {}

    print("siwakeApp IN " + str(Qnum))

    img_num = 0
    f_basename = ""

    siwake_win = tkinter.Tk()
    siwake_win.title("採点画面")
    siwake_win.geometry("1000x800")
    siwake_frame = tkinter.Frame(siwake_win)
    siwake_frame.grid(column=0, row=0)
    button_siwake_frame = tkinter.Frame(siwake_win)
    button_siwake_frame.grid(
        column=1, row=0, sticky=tkinter.W + tkinter.E + tkinter.N + tkinter.S)

    # キャンバス描画設定
    image_canvas = tkinter.Canvas(siwake_frame,
                                  bg="green",
                                  width=640,
                                  height=480)

    image_canvas.pack(expand=True, fill="both")

    # 仕分け結果表示
    assort_t_var = tkinter.StringVar(siwake_frame)
    assort_t_var.set("1 ~ 9のキーで点数を入力してください\n[space]で採点をskipします")
    assort_label = tkinter.Label(
        siwake_frame, textvariable=assort_t_var, font=("Meiryo UI", 30), bg="white",  relief="sunken")
    assort_label.pack()

    # ファイル名ラベル描画設定
    tex_var = tkinter.StringVar(siwake_frame)
    tex_var.set("ファイル名")

    lbl = tkinter.Label(siwake_frame, textvariable=tex_var,
                        font=("Meiryo UI", 20))
    lbl["foreground"] = "gray"
    lbl.pack()

    # 右左キーで画像送りする動作設定
    siwake_win.bind("<Key-Right>", next_img)
    siwake_win.bind("<Return>", next_img)
    siwake_win.bind("<space>", next_img)
    siwake_win.bind("<BackSpace>", prev_img)
    siwake_win.bind("<Key-Left>", prev_img)
    siwake_win.bind("<Delete>", prev_img)
    # 「Ctrl」+「P」で画像表示
    siwake_win.bind("<Control-Key-p>", image_show)

    # 数字キーで仕分け対象設定
    siwake_win.bind("<Key>", file_assort)

    # 仕分け実行ボタン
    assort_btn = tkinter.Button(
        button_siwake_frame, text="採点実行",  height=3, width=15)
    assort_btn.bind("<Button-1>", assort_go)

    # ファイル名ラベル描画設定
    saitenCount = tkinter.StringVar(button_siwake_frame)
    saitenCount.set("")
    ikutsuLb = tkinter.Label(
        button_siwake_frame, textvariable=saitenCount, font=("Meiryo UI", 20))
    ikutsuLb.pack(side=tkinter.TOP)

    exit_button = tkinter.Button(
        button_siwake_frame, text="トップに戻る\n保存はされません", height=3, width=15,  command=exit_siwake)
    exit_button.pack()

    # backfigB = tkinter.Label(siwake_frame, text="←前へ\nキーボードの←ボタン", font=(
    #     "Meiryo UI", 20)).pack(side=tkinter.LEFT, expand=TRUE)
    # nextfigB = tkinter.Label(siwake_frame, text="次へ→\nキーボードの→ボタン", font=(
    #     "Meiryo UI", 20)).pack(side=tkinter.RIGHT, expand=TRUE)

    ## 禁則処理のゾーン-------------
    setumeiBun1 = tkinter.Label(button_siwake_frame , text = "入力可能な点数にチェックをつけてください。" ).pack(side = tkinter.TOP)
    setumeiBun2 = tkinter.Label(button_siwake_frame , text = "誤った数字キーを押すのを防ぎます。").pack(side = tkinter.TOP)
    chkfonts = ("Meiryo UI", 10)
    cbln0 = tkinter.BooleanVar(master = siwake_win)
    chk0 = tkinter.Checkbutton(  master = button_siwake_frame,  variable=cbln0 ,text='0' , font = chkfonts).pack(side = tkinter.TOP)
    cbln1 = tkinter.BooleanVar(master = siwake_win)
    chk1 = tkinter.Checkbutton( variable=cbln1 , master = button_siwake_frame, text='1' , font = chkfonts).pack(side = tkinter.TOP)
    cbln2 = tkinter.BooleanVar(master = siwake_win)
    chk2 = tkinter.Checkbutton( variable=cbln2 , master = button_siwake_frame, text='2' , font = chkfonts).pack(side = tkinter.TOP)
    cbln3 = tkinter.BooleanVar(master = siwake_win)
    chk3 = tkinter.Checkbutton( variable=cbln3 , master = button_siwake_frame, text='3' , font = chkfonts).pack(side = tkinter.TOP)
    cbln4 = tkinter.BooleanVar(master = siwake_win)
    chk4 = tkinter.Checkbutton( variable=cbln4 , master = button_siwake_frame, text='4' , font = chkfonts).pack(side = tkinter.TOP)
    cbln5 = tkinter.BooleanVar(master = siwake_win)
    chk5 = tkinter.Checkbutton( variable=cbln5 , master = button_siwake_frame, text='5' , font = chkfonts).pack(side = tkinter.TOP)
    cbln6 = tkinter.BooleanVar(master = siwake_win)
    chk6 = tkinter.Checkbutton( variable=cbln6 , master = button_siwake_frame, text='6' , font = chkfonts).pack(side = tkinter.TOP)
    cbln7 = tkinter.BooleanVar(master = siwake_win)
    chk7 = tkinter.Checkbutton( variable=cbln7 , master = button_siwake_frame, text='7' , font = chkfonts).pack(side = tkinter.TOP)
    cbln8 = tkinter.BooleanVar(master = siwake_win)
    chk8 = tkinter.Checkbutton( variable=cbln8 , master = button_siwake_frame, text='8' , font = chkfonts).pack(side = tkinter.TOP)
    cbln9 = tkinter.BooleanVar(master = siwake_win)
    chk9 = tkinter.Checkbutton( variable=cbln9 , master = button_siwake_frame, text='9' , font = chkfonts).pack(side = tkinter.TOP)


    # 読み込みボタン描画設定
    load_file(Qnum)

    siwake_win.mainloop


def output_name_sh():

    # 定数設定
    SHEET_TITLE = '採点シート'  # シート名の設定
    RESULT_FILE_NAME = './setting/saiten.xlsx'  # 結果を保存するファイル名

    # 変数
    max_height = []  # 各行の画像の高さの最大値を保持

    def get_file_names(set_dir_name):
        """
        ディレクトリ内のファイル名取得（ファイル名のみの一覧を取得）
        """
        file_names = os.listdir(set_dir_name)
        temp_full_file_names = [os.path.join(set_dir_name, file_name) for file_name in file_names if os.path.isfile(
            os.path.join(set_dir_name, file_name))]  # ファイルかどうかを判定
        return temp_full_file_names

    def attach_img(target_full_file_names, set_column_idx, set_dir_name):
        """
        画像を呼び出して、Excelに貼り付け
        """
        set_row_idx = 1
        column_letter = "B"
        # 各列の1行目に、貼り付ける画像があるディレクトリ名を入力
        ws.cell(row=1, column=set_column_idx).value = "画像"
        ws.cell(row=1, column=1).value = "ファイル名"  # ファイル名
        ws.cell(row=1, column=3).value = "生徒番号"  # 出席番号
        ws.cell(row=1, column=4).value = "名前"  # 名前
        max_width = 0  # 画像の幅の最大値を保持するための変数
        target_full_file_names.sort()  # ファイル名でソート
        for target_file in target_full_file_names:
            p = pathlib.Path(target_file)
            target_file = p.resolve()
            if imghdr.what(target_file) != None:  # 画像ファイルかどうかの判定
                img = openpyxl.drawing.image.Image(target_file)
                #print('[' + column_letter + '][' + str(set_row_idx+1) + ']' + target_file + 'を貼り付け')

                # 画像のサイズを取得して、セルの大きさを合わせる（画像同士が重ならないようにするため）
                size_img = Image.open(target_file)
                width, height = size_img.size
                if max_width < width:
                    max_width = width
                # 配列「max_height」において、「set_row_idx」番目の要素が存在しなければ、挿入
                if not max_height[set_row_idx-1:set_row_idx]:
                    max_height.insert(set_row_idx-1, height)
                if max_height[set_row_idx-1] < height:
                    max_height[set_row_idx-1] = height
                ws.row_dimensions[set_row_idx + 1].height = max_height[set_row_idx-1] * 0.75
                ws.column_dimensions[column_letter].width = int(max_width) * 0.13

                # セルの行列番号から、そのセルの番地を取得
                cell_address = ws.cell(
                    row=set_row_idx + 1, column=set_column_idx).coordinate
                img.anchor = cell_address
                ws.add_image(img)  # シートに画像貼り付け
                ws.cell(row=set_row_idx + 1,
                        column=1).value = os.path.basename(target_file)

            set_row_idx += 1

    # ワークブック設定
    wb = openpyxl.Workbook()
    ws = wb.worksheets[0]  # 1番目のシートを編集対象にする
    ws.title = SHEET_TITLE  # 1番目のシートに名前を設定

    # 貼り付ける画像を置いておくルートディレクトリ内のディレクトリ名を再帰的に取得
    dir_name = "./setting/output/name"

    column_idx = 2

    f_names = get_file_names(dir_name)  # ファイル名取得
    attach_img(f_names, column_idx, dir_name)  # 画像貼り付け設定

    # ファイルへの書き込み
    wb.save(RESULT_FILE_NAME)


def writeImg():
    def readCSV_loc():
        # もしcsvが無ければ、全部止める
        if os.path.isfile("./setting/trimData.csv") == False:
            return 0
        else:
            with open('./setting/trimData.csv') as f:
                reader = csv.reader(f)
                data = [row for row in reader]
                data.pop(0)
                return data

    def getPoint(figName, qName):
        qCol = int(qName[-3:]) + 3
        xlPath = "./setting/saiten.xlsx"
        wb = openpyxl.load_workbook(xlPath)
        ws = wb["採点シート"]
        list = [[0 for i in range(ws.max_column)] for j in range(ws.max_row)]
        for x in range(0, ws.max_row):
            for y in range(0, ws.max_column):
                list[x][y] = ws.cell(row=x+1, column=y+1).value
        for row in list:
            if figName in row:
                result = True
                break
        result
        if not str(row[qCol]) == "未":
            return str(row[qCol])
        else:
            return "?"

    def pil2cv(image):
        ''' PIL型 -> OpenCV型 '''
        new_image = np.array(image, dtype=np.uint8)
        if new_image.ndim == 2:  # モノクロ
            pass
        elif new_image.shape[2] == 3:  # カラー
            new_image = cv2.cvtColor(new_image, cv2.COLOR_RGB2BGR)
        elif new_image.shape[2] == 4:  # 透過
            new_image = cv2.cvtColor(new_image, cv2.COLOR_RGBA2BGRA)
        return new_image


    files = glob.glob("./setting/input/*")
    data = readCSV_loc()
    img = Image.open(files[0])
    draw = ImageDraw.Draw(img)  # ImageDrawオブジェクトを作成
    if data:
        try:
            shutil.rmtree("./setting/kaitoYousi")
        except:
            print("kaitoYousiフォルダがまだありません、今から作ります。")
        os.makedirs("./setting/kaitoYousi", exist_ok=True)
        nmtitle, nmleft, nmtop, nmright, nmbottom = data.pop(0)
        h = abs(int(nmbottom) - int(nmtop))
        w = abs(int(nmright) - int(nmleft))
        if nmtitle == "name":
            if h >= w:
                mojiSize = int(w/2)
            else:
                mojiSize = int(h/2)
            try:
                font = ImageFont.truetype("arial.ttf", int(
                    mojiSize))  # フォントを指定、64はサイズでピクセル単位
            except:
                font = ImageFont.truetype("AppleGothic.ttf", int(
                    mojiSize))  # フォントを指定、64はサイズでピクセル単位
    for i,f in enumerate(files):
        img = Image.open(f)
        draw = ImageDraw.Draw(img)  # ImageDrawオブジェクトを作成
        print(os.path.basename(f) + "を採点します。")
        sumVal = 0
        for pos in data:
            title, left, top, right, bottom = pos
            figName = os.path.basename(f)
            qName = title
            text = getPoint(figName, qName)
            try:
                sumVal = sumVal + int(text)
            except:
                sumVal = sumVal
            # 画像
            draw.text((int(int(right) - mojiSize/2), int(top)),
                          text, font=font, fill="red")
            draw.rectangle((int(int(right) - mojiSize/2), int(top),
                               int(int(right) + mojiSize/2), int(top) + mojiSize),  outline="red")
        # 合計点を描画
        text = str(sumVal)
        draw.text((int(int(nmright) - mojiSize/2), int(nmtop)), text, font=font, fill="red")
        draw.rectangle((int(int(nmright) - mojiSize * 1.5), int(nmtop) + mojiSize,
                            int(int(nmright) + mojiSize * 1.5), int(nmtop) + mojiSize),  outline="red")
        df_zahyo = pd.read_csv("./setting/trimData.csv", index_col=0)
        daimon_list = df_zahyo.index[1:]
        df_zahyo = df_zahyo.T
        # 画像を読み込む
        img = pil2cv(img)
        jpg=f.split("ut")[1]
        # 問題番号リストで回す
        for daimon in daimon_list:
            if os.path.exists("./setting/output/"+daimon+"/mohan.png"):
                folda_max=os.listdir("./setting/output/"+daimon)[-2]
                shutil.move("./setting/output/"+daimon+"/mohan.png", "./setting/output/"+daimon+"/"+ folda_max +"/mohan.png")
            # 問題番号の座標を取得
            x_s,y_s,x_g,y_g=df_zahyo[daimon]
            x= round(x_s+(x_g-x_s)/2)
            y=round(y_s+(y_g-y_s)/2)
            # 大きさによって〇のサイズを変える
            if x_g-x_s < y_g-y_s:
                size = (x_g-x_s)/3
            elif y_g-y_s < x_g-x_s:
                size = (y_g-y_s)/3
            # 大問フォルダの中の配点フォルダ名を取得
            haiten_list=os.listdir("./setting/output/"+daimon)
            # 0点フォルダは最初
            haiten_0 = haiten_list[0]
            # 0点フォルダのpass
            img_path_0 = daimon +"/"+ haiten_0 + "/" + jpg
            # バツを付ける
            if os.path.exists("./setting/output/" + img_path_0):
                img = cv2.drawMarker(img, (x, y), (255, 0, 0), thickness=8, markerType=cv2.MARKER_TILTED_CROSS, markerSize=int(size))
            else:
                pass
        # 正解フォルダは最後
            haiten_cor = haiten_list[-1]
        # 正解フォルダのpass
            img_path_cor = daimon +"/"+ haiten_cor + "/" + jpg
        # 丸を付ける
            if os.path.exists("./setting/output/" + img_path_cor):
                img = cv2.circle(img, (x, y), int(size), (0, 0, 255), thickness=3, lineType=cv2.LINE_AA)
            else:
                pass
        # もし配点フォルダが２つなら、○×のみなのでpassする。
            if len(haiten_list) == 2:
                pass
            else:
                haiten_bubun = haiten_list[1:-1]
                for bubun in haiten_bubun:
                    img_path_bubun = daimon +"/"+ bubun + "/" + jpg
        # 三角を付ける
                    if os.path.exists("./setting/output/" + img_path_bubun):
                        img = cv2.drawMarker(img, (x, y), (0, 255, 0), thickness=3, markerType=cv2.MARKER_TRIANGLE_UP, markerSize=int(size))
                    else:
                        pass
        # セーブする
        cv2.imwrite("./setting/kaitoYousi/"+ jpg, img)


def writeImg_kanten(ginou, shikou):
    def pil2cv(image):
        ''' PIL型 -> OpenCV型 '''
        new_image = np.array(image, dtype=np.uint8)
        if new_image.ndim == 2:  # モノクロ
            pass
        elif new_image.shape[2] == 3:  # カラー
            new_image = cv2.cvtColor(new_image, cv2.COLOR_RGB2BGR)
        elif new_image.shape[2] == 4:  # 透過
            new_image = cv2.cvtColor(new_image, cv2.COLOR_RGBA2BGRA)
        return new_image
    
    def initDir2():
        os.makedirs("./setting/input", exist_ok=True)
        os.makedirs("./setting/output", exist_ok=True)
        f = open('setting/ini2.csv', 'w')  # 既存でないファイル名を作成してください
        writer = csv.writer(f, lineterminator='\n')  # 行末は改行
        writer.writerow(["tag", "start_x", "start_y", "end_x", "end_y"])
        f.close()

    def readCSV_loc():
        # もしcsvが無ければ、全部止める
        if os.path.isfile("./setting/trimData.csv") == False:
            return 0
        else:
            with open('./setting/trimData.csv') as f:
                reader = csv.reader(f)
                data = [row for row in reader]
                data.pop(0)
                return data    

    def getPoint(figName, qName):
        qCol = int(qName[-3:]) + 3
        xlPath = "./setting/saiten.xlsx"
        wb = openpyxl.load_workbook(xlPath)
        ws = wb["採点シート"]
        list = [[0 for i in range(ws.max_column)] for j in range(ws.max_row)]
        for x in range(0, ws.max_row):
            for y in range(0, ws.max_column):
                list[x][y] = ws.cell(row=x+1, column=y+1).value
            for row in list:
                if figName in row:
                    result = True
                    break
            if not str(row[qCol]) == "未":
                return str(row[qCol])
            else:
                return "?"


    def kanten_matrix(fn, n):
        # ファイルを開いてすべての行をリストで取得する
        if os.path.isfile("./setting/trimData.csv") == False:
            return 0
        else:
            with open('./setting/kanten_matrix.csv') as f:
                reader = csv.reader(f)
                data = [row for row in reader]
                data.pop(0)
                return data  

    def readCSV_kanten():
            # もしcsvが無ければ、全部止める
        if os.path.isfile("./setting/kanten.csv") == False:
            return 0
        else:
            with open('./setting/kanten.csv') as f:
                reader = csv.reader(f)
                data_k = [row for row in reader]
                data_k.pop(0)
                return data_k 
    def summary_table(df_shomon, ginou, shikou):
        ginou_list = df_shomon.iloc[:,ginou].sum(axis="columns").tolist()
        shikou_list = df_shomon.iloc[:,shikou].sum(axis="columns").tolist()
        kanten = pd.DataFrame(data=[ginou_list,shikou_list], index=["ginou","shikou"]).T
        nyuryoku=pd.read_csv("./seitoPDF_CSV/seito_meibo.csv")
        nyuryoku["合計"] = df_shomon.iloc[:,1:].sum(axis="columns")
        nyuryoku["技能"] = kanten["ginou"]
        nyuryoku["思考"] = kanten["shikou"]
        kanten.to_csv("./setting/kanten.csv",index=False)
        nyuryoku.to_excel("./kousa_nyuryoku.xlsx", index=False, encoding="SHIFT-JIS")
        print("観点別ファイル(kanten.csv)を出力しました。採点の集計を行ってください。入力用ファイル（kousa_nyuryoku.xlsx）も出力しました。素点をコピペしてください。")
        df_summary = pd.concat([nyuryoku, df_shomon.iloc[:,1:]],axis=1)
        df_summary.to_csv("./seitoPDF_CSV/kousa_summary.csv", index=False, encoding="SHIFT-JIS")
        return df_summary
    
    # 観点別をかく場所を選択する。
    initDir2()
    Kantenbetu()
    # 観点別の点数を色分けするためにリスト化する。
    dfq = pd.DataFrame()
    dfq["index"]=ginou
    dfq["kanten"]="ginou"
    dfg = pd.DataFrame()
    dfg["index"]=shikou
    dfg["kanten"]="shikou"
    df_kanten = pd.concat((dfq, dfg),axis=0).sort_values('index')
    kanten_list = df_kanten["kanten"]
    # エクセルデータから、観点別入りデータを追加する。
    file_name="./setting/saiten.xlsx"
    df_saiten=pd.read_excel(file_name)
    df_shomon = df_saiten.iloc[:-2, :].drop(['画像',"生徒番号","名前"],axis=1)
    df_shomon = df_shomon.replace("未", 0 )
    # 観点別入りデータを出力
    files = glob.glob("./setting/input/*")
    data = readCSV_loc()
    img = Image.open(files[0])
    draw = ImageDraw.Draw(img)  # ImageDrawオブジェクトを作成
    if data:
        try:
            shutil.rmtree("./setting/kaitoYousi")
        except:
            print("kaitoYousiフォルダがまだありません、今から作ります。")
        os.makedirs("./setting/kaitoYousi", exist_ok=True)
        nmtitle, nmleft, nmtop, nmright, nmbottom = data.pop(0)
        h = abs(int(nmbottom) - int(nmtop))
        w = abs(int(nmright) - int(nmleft))
        if nmtitle == "name":
            if h >= w:
                mojiSize = int(w/2)
            else:
                mojiSize = int(h/2)
            try:
                font = ImageFont.truetype("arial.ttf", int(
                    mojiSize))  # フォントを指定、64はサイズでピクセル単位
            except:
                font = ImageFont.truetype("AppleGothic.ttf", int(
                    mojiSize))  # フォントを指定、64はサイズでピクセル単位
    for i,f in enumerate(files):
        img = Image.open(f)
        draw = ImageDraw.Draw(img)  # ImageDrawオブジェクトを作成
        print(os.path.basename(f) + "を採点します。")
        sumVal = 0
        for pos, kan in zip(data, kanten_list):
            title, left, top, right, bottom = pos
            figName = os.path.basename(f)
            qName = title
            qCol = int(qName[-3:]) + 3
            xlPath = "./setting/saiten.xlsx"
            wb = openpyxl.load_workbook(xlPath)
            ws = wb["採点シート"]
            list = [[0 for i in range(ws.max_column)] for j in range(ws.max_row)]
            for x in range(0, ws.max_row):
                for y in range(0, ws.max_column):
                    list[x][y] = ws.cell(row=x+1, column=y+1).value
                for row in list:
                    if figName in row:
                        result = True
                        break
                if not str(row[qCol]) == "未":
                    text=str(row[qCol])
                else:
                    text="?"
        #         text = getPoint(figName, qName)
            if kan=="ginou":
                col = (200, 200, 64)
            elif kan=="shikou":
                col = "blue"
            # 画像
            draw.text( (int(int(right) - mojiSize/2), int(int(top) - int(mojiSize/10))), text, font=font, fill=col)
            draw.rectangle((int(int(right) - mojiSize/2), int(top), int(int(right)), int(top) + mojiSize*0.8),  outline="red")

        #観点別点数を描画
        title_g, left_g, top_g, right_g, bottom_g = kanten_matrix(img, 3)[0]
        title_s, left_s, top_s, right_s, bottom_s = kanten_matrix(img, 3)[1]
        data_kanten = readCSV_kanten()
        ginou, shikou = data_kanten[i]
        draw.text((int(int(right_g) - (int(right_g) - int(left_g))/1.5), int(int(bottom_g) - (int(bottom_g) - int(top_g))/1.5)), ginou, font=font, fill=(200, 200, 64))
        draw.text((int(int(right_s) - (int(right_s) - int(left_s))/1.5), int(int(bottom_s) - (int(bottom_s) - int(top_s))/1.5)), shikou, font=font, fill="blue")
        # 合計点を入力
        total_score = df_shomon.iloc[:,1:].sum(axis="columns")[i]
        sumtitle, sumleft, sumtop, sumright, sumbottom = kanten_matrix(img, 3)[2]
        try:
            font_sum = ImageFont.truetype("arial.ttf", int(
                            mojiSize*1.5))  # フォントを指定、64はサイズでピクセル単位
        except:
            font_sum = ImageFont.truetype("AppleGothic.ttf", int(
                            mojiSize*1.5))  # フォントを指定、64はサイズでピクセル単位
        draw.text((int(int(sumright)- (int(sumright) - int(sumleft))/1.5), int(int(sumbottom) - (int(sumbottom) - int(sumtop))/1.5)),
                  str(total_score), font = font_sum, fill="red")
        
        df_zahyo = pd.read_csv("./setting/trimData.csv", index_col=0)
        daimon_list = df_zahyo.index[1:]
        df_zahyo = df_zahyo.T
        # 画像を読み込む
        img = pil2cv(img)
        jpg=f.split("ut")[1]
        # 問題番号リストで回す
        for daimon in daimon_list:
            if os.path.exists("./setting/output/"+daimon+"/mohan.png"):
                folda_max=os.listdir("./setting/output/"+daimon)[-2]
                shutil.move("./setting/output/"+daimon+"/mohan.png", "./setting/output/"+daimon+"/"+ folda_max +"/mohan.png")
            # 問題番号の座標を取得
            x_s,y_s,x_g,y_g=df_zahyo[daimon]
            x= round(x_s+(x_g-x_s)/2)
            y=round(y_s+(y_g-y_s)/2)
            # 大きさによって〇のサイズを変える
            if x_g-x_s < y_g-y_s:
                size = (x_g-x_s)/3
            elif y_g-y_s < x_g-x_s:
                size = (y_g-y_s)/3
            # 大問フォルダの中の配点フォルダ名を取得
            haiten_list=os.listdir("./setting/output/"+daimon)
            # 0点フォルダは最初
            haiten_0 = haiten_list[0]
            # 0点フォルダのpass
            img_path_0 = daimon +"/"+ haiten_0 + "/" + jpg
            # バツを付ける
            if os.path.exists("./setting/output/" + img_path_0):
                img = cv2.drawMarker(img, (x, y), (255, 0, 0), thickness=8, markerType=cv2.MARKER_TILTED_CROSS, markerSize=int(size))
            else:
                pass
        # 正解フォルダは最後
            haiten_cor = haiten_list[-1]
        # 正解フォルダのpass
            img_path_cor = daimon +"/"+ haiten_cor + "/" + jpg
        # 丸を付ける
            if os.path.exists("./setting/output/" + img_path_cor):
                img = cv2.circle(img, (x, y), int(size), (0, 0, 255), thickness=3, lineType=cv2.LINE_AA)
            else:
                pass
        # もし配点フォルダが２つなら、○×のみなのでpassする。
            if len(haiten_list) == 2:
                pass
            else:
                haiten_bubun = haiten_list[1:-1]
                for bubun in haiten_bubun:
                    img_path_bubun = daimon +"/"+ bubun + "/" + jpg
        # 三角を付ける
                    if os.path.exists("./setting/output/" + img_path_bubun):
                        img = cv2.drawMarker(img, (x, y), (0, 255, 0), thickness=3, markerType=cv2.MARKER_TRIANGLE_UP, markerSize=int(size))
                    else:
                        pass
        # セーブする
        cv2.imwrite("./setting/kaitoYousi/"+ jpg, img)
        print(os.path.basename(f) + "の採点完了しました。")
        
def resize_pic(pic1):
    #調整後サイズを指定(横幅、縦高さ)
    size=(28,28)
    base_pic=np.zeros((size[1],size[0]),np.uint8)
#     pic1=cv2.cvtColor(pic1, cv2.COLOR_BGR2GRAY)
    h,w=pic1.shape[:2]
    ash=size[1]/h
    asw=size[0]/w
    if asw<ash:
        sizeas=(int(w*asw),int(h*asw))
    else:
        sizeas=(int(w*ash),int(h*ash))
    pic1 = cv2.resize(pic1,dsize=sizeas)
    base_pic[int(size[1]/2-sizeas[1]/2):int(size[1]/2+sizeas[1]/2),int(size[0]/2-sizeas[0]/2):int(size[0]/2+sizeas[0]/2)]=pic1
    return base_pic

def IMAGE_TO_CICLEorCROSS(BMP):
    image_width=28
    image_height=28
    color_setting=1
    # マルバツモデルを使う準備
    ci_cro = ["maru", "batu"]
    model = model_from_json(open('saiten/cnn_model.json', 'r').read())
    model.load_weights('saiten/cicle_cross.h5')
    img = cv2.imread(BMP, 0)
    img = resize_pic(pic1 = img)
    img = img.reshape(image_width, image_height, color_setting).astype('float32')/255 
    prediction = model.predict(np.array([img]))
    result = prediction[0]
    return ci_cro[result.argmax()]
        
def IMAGE_TO_NUM_SINGLE(BMP):
    image_width=28
    image_height=28
    color_setting=1
    number = ["0","1","2","3","4","5","6","7","8","9"]
    model= load_model('saiten/number.h5')
    img = cv2.imread(BMP, 0)
    img = resize_pic(pic1 = img)
    img = img.reshape(image_width, image_height, color_setting).astype('float32')/255 
    prediction = model.predict(np.array([img]))
    result = prediction[0]
    return number[result.argmax()]
        
def IMAGE_TO_PREDICT(BMP):
    image_width=28
    image_height=28
    color_setting=1
    trim_lists=[]
    num_lists = []
    img = cv2.imread(BMP)
    img_gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    # 軽く膨張処理
#     img_gray = cv2.dilate(img_gray, None, iterations = 1)
    contours, _ = cv2.findContours(img_gray, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    for rect in contours:
    # 検出できた個数
    #     print(cv2.contourArea(rect))
        if cv2.contourArea(rect) > 10:
    #余分なものを取り除いた個数
    #         print(cv2.contourArea(rect))
            x1, y1, w, h = cv2.boundingRect(rect)
            if x1!=0:
                x1=x1-1
            x2=x1+w+2
            if y1!=0:
                y1 = y1-1
            y2=y1+h+2
    # 高さが下すぎるところからはじまるもの（y1）、高さが上すぎるところで終わるもの（y2）は省く 
            if y1 < img_gray.shape[0]-10 and y2 > 10 and y2-y1 > 10:
                trim_lists.append([x1,x2,y1,y2])
    trim_lists.sort()
    for trim_list_i in trim_lists:
        x1,x2,y1,y2 = trim_list_i
        y3 = y2 if y1 > y2 else y1
        y4 = y1 if y2 < y1 else y2
        y1 = y3
        y2 = y4
        img = cv2.imread(BMP)
        img_gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    #軽く膨張処理
        img_test = img_gray[y1:y2, x1:x2] 
    # 以下、mnistのデータセットに合わせて加工した
        img_re = resize_pic(img_test)# これが保存される
        img_resh = img_re.reshape(image_width, image_height, color_setting).astype('float32')/255
        model= load_model('saiten/number.h5')
        prediction = model.predict(np.array([img_resh]))
        result = prediction[0]
        num = result.argmax()
    ####################### 結果の確認  
        num_lists.append(num)
        # 画像と結果の照らし合わせ
        pre_ans=""
        for num in num_lists:
            pre_ans += str(num)
    return pre_ans

def IMAGE_TO_LATEX(jpg):
    from pix2tex import cli as pix2tex
    from PIL import Image
    from munch import Munch
    arguments = Munch({'config': './config.yaml', 'checkpoint': "C:/Users/suish/Desktop/OCR/LaTeX-OCR/pix2tex/mixed_e20_step580.pth", 'no_cuda': True, 'no_resize': False})
    model = pix2tex.LatexOCR(arguments)
    img = Image.open(jpg)
    math = model(img)
    return math

def PRE_MATH_ANS():
    global answer_text_list
    global folda
    answer_text_list=[]
    img_path = "./deep/"+ folda + "/*.png"
    png_list = glob.glob(img_path)
    for png in png_list:
        print(f"{png}をさいてんするぜ")
        a=IMAGE_TO_LATEX(png)
        print(a)
        answer_text_list.append(a)
    all_saiten.destroy()
    summary_saiten(folda, answer_text_list)

def PRE_NUM_ANS():
    global answer_text_list
    global folda
    answer_text_list=[]
    img_path = "./deep/"+folda + "/off_image_test/*.bmp"
    BMP_list = glob.glob(img_path)
    for BMP in BMP_list:
        print(f"{BMP}をさいてんするぜ")
        a=IMAGE_TO_PREDICT(BMP)
        print(a)
        answer_text_list.append(a)
    all_saiten.destroy()
    summary_saiten(folda, answer_text_list)
    
def summary_saiten(folda, answer_text_list):
    global answer_list
    global all_saiten

    class ClassFrame(tk.Frame):
        def __init__(self, master, bg=None, width=None, height=None):
            super().__init__(master, bg=bg, width=width, height=height)


    class ClassLabelFrameTop(tk.LabelFrame):
        def __init__(self, master, text=None, pad_x=None, pad_y=None, bg=None):
            super().__init__(master, text=text, padx=pad_x, pady=pad_y, bg=bg)
            button1 = tk.Button(self, text="完了", bg=bg)
            button1.pack(side="left", padx=(10, 0))
            button2 = tk.Button(self, text="タブ2", bg=bg)
            button2.pack(side="left", padx=(10, 0))
            button3 = tk.Button(self, text="タブ3", bg=bg)
            button3.pack(side="left", padx=(10, 0))


    class ClassLabelFrameLeft(tk.LabelFrame):
        def __init__(self, master, text=None, pad_x=None, pad_y=None, bg=None):
            super().__init__(master, text=text, padx=pad_x, pady=pad_y, bg=bg)
            left_button1 = tk.Button(self, text="完了", bg=bg, command=btn_click)
            left_button1.pack(anchor=tk.NW, fill=tk.X, padx=(10, 10), pady=(0, 10))
#             left_button2 = tk.Button(self, text="○×採点", bg=bg, command=maru_or_batu)
#             left_button2.pack(anchor=tk.NW, fill=tk.X, padx=(10, 10), pady=(0, 10))
#             left_button3 = tk.Button(self, text="0～9採点", bg=bg, command=number_0to9)
#             left_button3.pack(anchor=tk.NW, fill=tk.X, padx=(10, 10), pady=(0, 10))
            left_button4 = tk.Button(self, text="2桁以上数字採点", bg=bg, command=PRE_NUM_ANS)
            left_button4.pack(anchor=tk.NW, fill=tk.X, padx=(10, 10), pady=(0, 10))
            left_button5 = tk.Button(self, text="数式採点", bg=bg, command=PRE_MATH_ANS)
            left_button5.pack(anchor=tk.NW, fill=tk.X, padx=(10, 10), pady=(0, 10))
            
            
    class ClassCanvas(tk.Canvas):
        def __init__(self, master, scroll_width, scroll_height, bg):
            super().__init__(master, bg=bg)
            # Scrollbarを生成してCanvasに配置処理
            bar_y = tk.Scrollbar(self, orient=tk.VERTICAL)
            bar_x = tk.Scrollbar(self, orient=tk.HORIZONTAL)
            bar_y.pack(side=tk.RIGHT, fill=tk.Y)
            bar_x.pack(side=tk.BOTTOM, fill=tk.X)
            bar_y.config(command=self.yview)
            bar_x.config(command=self.xview)
            self.config(yscrollcommand=bar_y.set, xscrollcommand=bar_x.set)
            # Canvasのスクロール範囲を設定
            self.config(scrollregion=(0, 0, scroll_width, scroll_height))

        # マウスホイールに対応
            bar_y.bind('<Up>', lambda e:self.yview_scroll(-1,tk.UNITS))
            bar_y.bind('<Down>', lambda e:self.yview_scroll(1,tk.UNITS))
            bar_y.bind('<MouseWheel>', lambda e:self.yview_scroll(-1*(1 if e.delta>0 else -1),tk.UNITS))
            bar_y.bind('<Enter>',lambda e:bar_y.focus_set())
            self.bind('<MouseWheel>', lambda e:self.yview_scroll(-1*(1 if e.delta>0 else -1),tk.UNITS))
            self.bind('<Enter>',lambda e:bar_y.focus_set())
            bar_x.bind('<MouseWheel>', lambda e:self.xview_scroll(-1*(1 if e.delta>0 else -1),tk.UNITS))
            bar_x.bind('<Enter>',lambda e:bar_x.focus_set())
            self.config(xscrollcommand=bar_x.set)
            self.config(yscrollcommand=bar_y.set)
    
    def txt_mohan_place(frame_big_right,y):
        txt_mohan = tk.Entry(frame_big_right, width=20)
        txt_mohan.place(x=5, y=y)
        txt_mohan_ten = tk.Entry(frame_big_right, width=5)
        txt_mohan_ten.place(x=140, y=y)
        return txt_mohan, txt_mohan_ten

    def file_move(mohan_ten, img):
        dir_path = "./setting/output/" +folda+"/"
        name = img.split("setting/output/" +folda+"/")[1]
        if not os.path.exists(dir_path+mohan_ten):
            os.mkdir(dir_path+mohan_ten)
        shutil.move(img, dir_path + mohan_ten + "/" + name)
        
    def btn_click():
        global txt
        global answer_list
# 解答として処理するものを５つ用意
        mohan = txt_mohan.get()
        mohan_ten = txt_mohan_ten.get()
        mohan_2 = txt_mohan_2.get()
        mohan_ten_2 = txt_mohan_ten_2.get()    
        mohan_3 = txt_mohan_3.get()
        mohan_ten_3 = txt_mohan_ten_3.get()    
        mohan_4 = txt_mohan_4.get()
        mohan_ten_4 = txt_mohan_ten_4.get()    
        mohan_5 = txt_mohan_5.get()
        mohan_ten_5 = txt_mohan_ten_5.get()
        mohan_6 = txt_mohan_6.get()
        mohan_ten_6 = txt_mohan_ten_6.get()
    # 入力された解答データの取得
        new_ans_list=[] 
        img_list = sorted(glob.glob("./setting/output/"+folda+"/*.png"))
        for i,txt in enumerate(answer_list):
            exec_command = 'txt_' + str(i) + '=' + "txt"
            exec(exec_command)
            answer = 'new_ans_list.append(txt_' + str(i) + '.get())'
            exec(answer)
        hold_ans_list=[]
    # allと入力されてたら一気に移動
        if mohan_6 =="all":
            dir_path = "./setting/output/" +folda+"/"
            img_list = glob.glob("./setting/output/"+folda+"/*.png")
            if not os.path.exists(dir_path+mohan_ten_6):
                os.mkdir(dir_path+mohan_ten_6)
            for img in img_list:
                if img == "./setting/output/"+folda+"/mohan.png":
                    print("模範解答は移動しません")
                    hold_ans_list.append("mohan")
                    continue
                name = img.split("setting/output/" +folda+"/")[1]
                shutil.move(img, dir_path + mohan_ten_6 + "/" + name)
    # 模範解答欄に入力された物事に移動      
        else:        
            for img_i,img in enumerate(img_list):
                if img == "./setting/output/"+folda+"/mohan.png":
                    print("模範解答は移動しません")
                    hold_ans_list.append(new_ans_list[img_i])
                    continue
                else:
                    if new_ans_list[img_i] =="" or new_ans_list[img_i] =="?":
                        print("空欄なので【 0 】を入れておきます")
                        hold_ans_list.append("0")
                    elif new_ans_list[img_i] == mohan:
                        print(f"「{mohan_ten}」点")
                        file_move(mohan_ten, img)
                    elif new_ans_list[img_i] == mohan_2:
                        print(f"「{mohan_ten_2}」点")
                        file_move(mohan_ten_2, img)
                    elif new_ans_list[img_i] == mohan_3:
                        print(f"「{mohan_ten}」点")
                        file_move(mohan_ten_3, img)
                    elif new_ans_list[img_i] == mohan_4:
                        print(f"「{mohan_ten}」点")
                        file_move(mohan_ten_4, img)
                    elif new_ans_list[img_i] == mohan_5:
                        print(f"「{mohan_ten}」点")
                        file_move(mohan_ten_5, img)
                    else:
                        print("入力は確認しましたが、移動はしません。解答を保存しておきます")
                        hold_ans_list.append(new_ans_list[img_i])
# 解答データの保存
        df_ans_summmary_fix = pd.read_csv('./seitoPDF_CSV/df_ans_summmary_fix.csv',index_col = 0)
        zan_file = glob.glob("./setting/output/" + folda+ "/*.png")
        im_file_name = [file.split(folda)[1] for file in zan_file]
        val = pd.Series(hold_ans_list, index=im_file_name)
        df_ans_summmary_fix[folda]=val
        df_ans_summmary_fix.to_csv('./seitoPDF_CSV/df_ans_summmary_fix.csv')
        messagebox.showinfo("移動終了", "画像の移動が終わりました。入力したものも保存しました。")
        all_saiten.destroy()
        list_ = df_ans_summmary_fix[[folda]].dropna()[folda].to_list()
        answer_text_list=[round(i) if type(i) is float else i for i in list_ ]
        summary_saiten(folda, answer_text_list)
    
    all_saiten = tk.Tk()
    all_saiten.geometry("1000x800") 
    bg_color = "snow"
    # 最上位のフレーム
    frame_big_left = ClassFrame(all_saiten, bg="green")
    frame_big_left.pack(side=tk.LEFT, expand=1, fill=tk.BOTH)
    
    # 採点するフォルダリスト
    # 採点する画像リスト
    img_list = sorted(glob.glob("./setting/output/" +folda+"/*.png"))
    # 模範解答でいろいろ設定を行う
    img_mohan = Image.open("./setting/output/" +folda+"/mohan.png")
    # 横の長さｗと縦の長さｈ
    w_mohan = img_mohan.size[0]
    h_mohan = img_mohan.size[1]

    # 画像の枚数×（リサイズした高さ＋１５０－リサイズした模範解答の高さ）÷４
    if w_mohan < 520:
        width_size=260
        img_mohan = img_mohan.resize((width_size, (h_mohan*width_size)//w_mohan))
        w_mohan = img_mohan.size[0]
        h_mohan = img_mohan.size[1]
        y_scrol = len(img_list)*(h_mohan + 150)//4
    else:
        width_size = 520
        img_mohan = img_mohan.resize((width_size, (h_mohan*width_size)//w_mohan))
        w_mohan = img_mohan.size[0]
        h_mohan = img_mohan.size[1]
        y_scrol = len(img_list)*(h_mohan + 150)//2
        width_size=260
        img_mohan = img_mohan.resize((width_size, (h_mohan*width_size)//w_mohan))
        w_mohan = img_mohan.size[0]
        h_mohan = img_mohan.size[1]

    # 右に模範解答を表示する
    frame_big_right = ClassFrame(all_saiten, bg="purple", width = w_mohan+10)
    frame_big_right.pack(side=tk.RIGHT, expand=0, fill=tk.Y)
    canvas_mohan = tk.Canvas(frame_big_right, bg = "purple", width = w_mohan, height = h_mohan) 
    canvas_mohan.place(x=5, y=5)
    img_tk_mohan = ImageTk.PhotoImage(img_mohan, master=canvas_mohan)
    canvas_mohan.create_image(0, 0, image=img_tk_mohan, anchor=tk.NW)

    txt_mohan   , txt_mohan_ten  = txt_mohan_place(frame_big_right,y=h_mohan + 15)
    txt_mohan_2 ,txt_mohan_ten_2 = txt_mohan_place(frame_big_right,y=h_mohan + 40)
    txt_mohan_3 ,txt_mohan_ten_3 = txt_mohan_place(frame_big_right,y=h_mohan + 65)
    txt_mohan_4 ,txt_mohan_ten_4 = txt_mohan_place(frame_big_right,y=h_mohan + 90)
    txt_mohan_5 ,txt_mohan_ten_5 = txt_mohan_place(frame_big_right,y=h_mohan + 115)
    txt_mohan_6 ,txt_mohan_ten_6 = txt_mohan_place(frame_big_right,y=h_mohan + 500)

    # スクロールの縦の長さをきめる.Canvasを生成
    scroll_max = {"width": 1000, "height": y_scrol + 200}
    canvas = ClassCanvas(frame_big_left, scroll_width=scroll_max["width"],scroll_height=scroll_max["height"], bg="green")
    canvas.place(x=0, y=0, relheight=1, relwidth=1)

    x,y=5,0
    imgs,answer_list=[],[]
    for i in range(len(img_list)):
        img = Image.open(img_list[i])
        w=img.size[0]
        h=img.size[1]
    # Frame Widgetを 生成
        frame = tk.Frame(canvas, width=160,height=30, relief=tk.RAISED, bg='green', bd=2)
        if w < 520:
            width_size=260
            img = img.resize((width_size, (h*width_size)//w))
            w = img.size[0]
            h = img.size[1]
            img = ImageTk.PhotoImage(img,master=canvas)
            imgs.append(img)
            canvas.create_image(x, y, image=img, anchor=tk.NW)
    # Frame Widgetを Canvas Widget上に配置
            canvas.create_window((x + 10, y + h + 20), window=frame, anchor=tk.NW)
            x += 270
            if (i+1) % 4 == 0:
                x = 5
                if h < 150:
                    y += 150
                else:
                    y += h + 150
        else:
            width_size=520
            img = img.resize((width_size, (h*width_size)//w))
            w = img.size[0]
            h = img.size[1]
            img = ImageTk.PhotoImage(img,master=canvas)
            imgs.append(img)
            canvas.create_image(x, y, image=img, anchor=tk.NW)
    # Frame Widgetを Canvas Widget上に配置
            canvas.create_window((x + 10, y + h + 20), window=frame, anchor=tk.NW)
            x += 540
            if (i+1) % 2 == 0:
                x = 5
                if h < 150:
                    y += 150
                else:
                    y += h + 150
        text = tk.StringVar().set("1")
        txt = tk.Entry(frame, width=20,bg = "white",textvariable=text)
        if not answer_text_list == []:
            txt.insert("1", answer_text_list[i])
        txt.place(x=15,y=2)
        txt.bind('<Up>', lambda e:canvas.yview_scroll(-1,tk.UNITS))
        txt.bind('<Down>', lambda e:canvas.yview_scroll(1,tk.UNITS))
        answer_list.append(txt)

    # 右側の大フレームの中に設置するラベル付き縦フレームのテスト
    label_frame_left_menu = ClassLabelFrameLeft(frame_big_right, text="メニュー", pad_y=3, bg=bg_color)
    label_frame_left_menu.place(x=10, y= h_mohan + 150, width=200)

    # 右側の大フレームの中に設置するlabelFrame
    # label_frame_top = ClassLabelFrameTop(frame_big_right, text="メニュー", pad_x=0, pad_y=5, bg=bg_color)
    # label_frame_top.place(x=10, y=500, width=200, height=50)
    # label_frame_top.lift()

    # 右側のFrameの中に設置するFrame
    # frame_test = ClassFrame(frame_big_left, bg="grey", width=300, height=100)
    # frame_test.place(x=0, y=0)
    # frame_test.pack(side=tk.RIGHT, anchor=tk.SW, expand=0, padx=(10, 17), pady=(0, 25))
    all_saiten.mainloop()
    
def mohan_chice_saiten():
    global mohan_divide_list
    global mohan_saiten
    global saiten_folda
    class ClassFrame(tk.Frame):
        def __init__(self, master, bg=None, width=None, height=None):
            super().__init__(master, bg=bg, width=width, height=height)


    class ClassLabelFrameLeft(tk.LabelFrame):
        def __init__(self, master, text=None, pad_x=None, pad_y=None, bg=None):
            super().__init__(master, text=text, padx=pad_x, pady=pad_y, bg=bg)
            left_button1 = tk.Button(self, text="自動判定", bg=bg, command=btn_click_saiten)
            left_button1.pack(anchor=tk.NW, fill=tk.X, padx=(10, 10), pady=(0, 10))
            
            
    class ClassCanvas(tk.Canvas):
        def __init__(self, master, scroll_width, scroll_height, bg):
            super().__init__(master, bg=bg)
            # Scrollbarを生成してCanvasに配置処理
            bar_y = tk.Scrollbar(self, orient=tk.VERTICAL)
            bar_x = tk.Scrollbar(self, orient=tk.HORIZONTAL)
            bar_y.pack(side=tk.RIGHT, fill=tk.Y)
            bar_x.pack(side=tk.BOTTOM, fill=tk.X)
            bar_y.config(command=self.yview)
            bar_x.config(command=self.xview)
            self.config(yscrollcommand=bar_y.set, xscrollcommand=bar_x.set)
            # Canvasのスクロール範囲を設定
            self.config(scrollregion=(0, 0, scroll_width, scroll_height))

        # マウスホイールに対応
            bar_y.bind('<Up>', lambda e:self.yview_scroll(-1,tk.UNITS))
            bar_y.bind('<Down>', lambda e:self.yview_scroll(1,tk.UNITS))
            bar_y.bind('<MouseWheel>', lambda e:self.yview_scroll(-1*(1 if e.delta>0 else -1),tk.UNITS))
            bar_y.bind('<Enter>',lambda e:bar_y.focus_set())
            self.bind('<MouseWheel>', lambda e:self.yview_scroll(-1*(1 if e.delta>0 else -1),tk.UNITS))
            self.bind('<Enter>',lambda e:bar_y.focus_set())
            bar_x.bind('<MouseWheel>', lambda e:self.xview_scroll(-1*(1 if e.delta>0 else -1),tk.UNITS))
            bar_x.bind('<Enter>',lambda e:bar_x.focus_set())
            self.config(xscrollcommand=bar_x.set)
            self.config(yscrollcommand=bar_y.set)
    

    def btn_click_saiten():
        global txt
        global mohan_divide_list
        global saiten_folda
        new_mohan_list=[]
        path="./"
        
        for i,txt in enumerate(mohan_divide_list):
            exec_command = 'txt_' + str(i) + '=' + "txt"
            exec(exec_command)
            answer = 'new_mohan_list.append(txt_' + str(i) + '.get())'
            exec(answer)
        folda_list = glob.glob("./deep/*")
        for m, folda in zip(new_mohan_list, saiten_folda):
#             folda = folda.split("\\")[1]
            print(f"{folda}の採点をしています。しばらくおまちください。")
            # 解答deepの読み込み
            ans_file_deep=glob.glob(path+"deep/"+folda +"/*.png")

            # 正解をリストの保存
            pred=[]
            if m== "1" or m== "2" or m=="4":
                img_path = path+"deep/"+folda + "/off_image_test/*.bmp"
                BMP_list = glob.glob(img_path)
                for BMP in BMP_list:
                    print(f"{BMP}をさいてんするぜ")
                    if m=="1":
                        try:
                            a=IMAGE_TO_NUM_SINGLE(BMP)
                        except:
                            a="?"
                    elif m== "2":
                        try:
                            a=IMAGE_TO_PREDICT(BMP)
                        except:
                            a="?"
                    elif m == "4":
                        try:
                            a=IMAGE_TO_CICLEorCROSS(BMP)
                        except:
                            a="?"
                    print(a)
                    pred.append(a)
                val = pd.Series(pred,index=[file.split(folda)[1] for file in ans_file_deep])
                df_ans_summmary[folda]=val  
                                     
            # 以下数式採点
            if m=="3":
                imgs = glob.glob(path+"deep/"+folda+"/*.png")
                for f in imgs:
                    print(f"{f}をさいてんするぜ")
                    try:
                        math = IMAGE_TO_LATEX(f)
                    except:
                        math = "?"
                    print(math)
                    pred.append(math)
                val = pd.Series(pred,index=[file.split(folda)[1] for file in ans_file_deep])
                df_ans_summmary[folda]=val

        #     if m== "5":
        #         ここにマーク採点
            else:
                print(f"{folda}は採点しません。")
            # 正解をまとめる
            print(f"{folda}の採点が完了しました。")
        print("全て終えたので保存しておきます。")
        df_ans_summmary.to_csv(path+'seitoPDF_CSV/df_ans_summmary.csv')
        df_ans_summmary.to_csv(path+'seitoPDF_CSV/df_ans_summmary_fix.csv')        
    path ="./"
    #横カラム
    folda_list = glob.glob(path+"setting/output/*")
    
    folda_colum = [col.split("t/")[1] for col in folda_list[1:]]
    #縦インデックス
    file_input = glob.glob(path+"setting/input/*.png")
    file_index = [file.split("ut")[1] for file in file_input]
    
    if not os.path.exists('./seitoPDF_CSV/df_ans_summmary.csv'):
        df_ans_summmary = pd.DataFrame(index = file_index, columns=folda_colum)
    else:
        ret = messagebox.askyesno('確認', 'すでに一度実行されています。上書きしても良いですか？')
        if ret == True:           
            df_ans_summmary = pd.DataFrame(index = file_index, columns=folda_colum)
        else:
            messagebox.showinfo(
                '終了してください', 'df_ans_summaryを上書き、新たにつくりません。画面は表示しますが、終了してください。')

    mohan_saiten = tk.Tk()
    mohan_saiten.geometry("1000x800") 
    bg_color = "snow"
    folda_list = sorted(glob.glob("./setting/output/*")[1:])
    mohan_img_list = []
    saiten_folda=[]
    for folda in folda_list:
        folda = folda.split("t/")[1]
        deep_folda_list =glob.glob(path+"deep/"+folda +"/*.png")
        output_folda_list =glob.glob(path+"setting/output/"+folda +"/*.png")
#         if len(deep_folda_list) == len(output_folda_list):
        mohan_img_list.append(path+"setting/output/"+ folda +"/mohan.png" )
        saiten_folda.append(folda)
#         else:
#             print(f"DEEPフォルダの画像枚数{len(deep_folda_list)}")
#             print(f"outputフォルダの画像枚数{len(output_folda_list)}")
#             continue
    # 右に模範解答を表示する
    frame_big_right = ClassFrame(mohan_saiten, bg="purple", width = 250)
    frame_big_right.pack(side=tk.RIGHT, expand=0, fill=tk.Y)
    right_lavel1 = tk.Label(frame_big_right,text='【1】採点（0～9）', foreground='linen', background='lightslategray')
    right_lavel1.place(x=10, y= 5, width=200)
    right_lavel1 = tk.Label(frame_big_right,text='【2】数字採点（2桁以上）', foreground='linen', background='lightslategray')
    right_lavel1.place(x=10, y= 35, width=200)
    right_lavel1 = tk.Label(frame_big_right,text='【3】数式採点', foreground='linen', background='lightslategray')
    right_lavel1.place(x=10, y= 65, width=200)
    right_lavel1 = tk.Label(frame_big_right,text='【4】○×採点', foreground='linen', background='lightslategray')
    right_lavel1.place(x=10, y= 95, width=200)
    right_lavel1 = tk.Label(frame_big_right,text='【5】マークシート採点', foreground='linen', background='lightslategray')
    right_lavel1.place(x=10, y= 125, width=200)
    right_lavel1 = tk.Label(frame_big_right,text='【0】フォルダを削除する', foreground='linen', background='lightslategray')
    right_lavel1.place(x=10, y= 155, width=200)
    # 右側の大フレームの中に設置するラベル付き縦フレームのテスト
    label_frame_left_menu = ClassLabelFrameLeft(frame_big_right, text="メニュー", pad_y=3, bg=bg_color)
    label_frame_left_menu.place(x=10, y= 185, width=200)
    
    y_scrol = len(mohan_img_list)*(320)//4
    
    
    # 最上位のフレーム
    frame_big_left = ClassFrame(mohan_saiten, bg="green")
    frame_big_left.pack(side=tk.LEFT, expand=1, fill=tk.BOTH)
    # スクロールの縦の長さをきめる.Canvasを生成
    scroll_max = {"width": 1000, "height": y_scrol + 200}
    canvas = ClassCanvas(frame_big_left, scroll_width=scroll_max["width"],scroll_height=scroll_max["height"], bg="green")
    canvas.place(x=0, y=0, relheight=1, relwidth=1)

    x,y=5,0
    
    imgs, mohan_divide_list=[],[]
    for i in range(len(mohan_img_list)):
        try:
            img = Image.open(mohan_img_list[i])
        except:
            print("見当たらないので飛ばします")
            continue
        w=img.size[0]
        h=img.size[1]
    # Frame Widgetを 生成
        frame = tk.Frame(canvas, width=160,height=30, relief=tk.RAISED, bg='green', bd=2)
        width_size=260
        img = img.resize((width_size, (h*width_size)//w))
        if h > 80:
            img = img.resize((width_size, 80))
        w = img.size[0]
        h = img.size[1]
        img = ImageTk.PhotoImage(img,master=canvas)
        imgs.append(img)
        canvas.create_image(x, y, image=img, anchor=tk.NW)
    # Frame Widgetを Canvas Widget上に配置
        canvas.create_window((x + 10, y + h + 20), window=frame, anchor=tk.NW)
        x += 270
        if (i+1) % 4 == 0:
            x = 5
            if h < 150:
                y += 150
            else:
                y += h + 150

        text = tk.StringVar().set("1")
        txt = tk.Entry(frame, width=20,bg = "white",textvariable=text)
#         if not mohan_text_list == []:
#             txt.insert("1", mohan_text_list[i])
        txt.place(x=15,y=2)
        txt.bind('<Up>', lambda e:canvas.yview_scroll(-1,tk.UNITS))
        txt.bind('<Down>', lambda e:canvas.yview_scroll(1,tk.UNITS))
        mohan_divide_list.append(txt)

    mohan_saiten.mainloop()
     

          
    


def top_activate():

    val = 0.4
    fifwid = 500
    fifhet = 400
    global root
    global topimg
    global topfig

    global top_frame
    top_frame = tkinter.Frame(root, bg="white")
    top_frame.pack()
    fig_frame = tkinter.Frame(top_frame, width=fifwid, height=fifhet)
    fig_frame.grid(column=0, row=0)

    try:
        topimg = Image.open(resource_path("C:/top1.png"))
        topimg = topimg.resize(
            (int(topimg.width * val), int(topimg.height * val)), 0)
        topfig = ImageTk.PhotoImage(topimg, master=root)
        canvas_top = tkinter.Canvas(
            bg="white", master=fig_frame, width=fifwid + 30, height=fifhet, highlightthickness=0)
        canvas_top.place(x=0, y=0)
        canvas_top.create_image(0, 0, image=topfig, anchor=tkinter.NW)
        canvas_top.pack()
    except:
        pass

    button_frame = tkinter.Frame(top_frame, bg="white", highlightthickness=0)
    button_frame.grid(column=1, row=0, sticky=tkinter.W +
                      tkinter.E + tkinter.N + tkinter.S)

    exBool = True
    botWid = 20

    # infoB = tkinter.Button(
    #     button_frame, text="はじめに", command=info, width=botWid, height=2, highlightthickness=0).pack(expand=exBool)

    # initB = tkinter.Button(
    #     button_frame, text="初期設定をする", command=setting_ck, width=botWid, height=2, highlightthickness=0).pack(expand=exBool)

#     GiriGoB = tkinter.Button(
#         button_frame, text="名前のみをきる", command=input_ck, width=botWid, height=2, highlightthickness=0).pack(expand=exBool)

    initB = tkinter.Button(
        button_frame, text="AI採点する問題を選んで採点", command=mohan_chice_saiten, width=botWid, height=2, highlightthickness=0).pack(expand=exBool)

    saitenB = tkinter.Button(
        button_frame, text="斬った画像を採点する", command=saitenSelect, width=botWid, height=2, highlightthickness=0).pack(expand=exBool)

#     outxlsxB = tkinter.Button(
#         button_frame, text="Excelに出力", command=outputXlsx, width=botWid, height=2, highlightthickness=0).pack(expand=exBool)

    writeImgB = tkinter.Button(
        button_frame, text="採点済み画像を出力", command=writeImg, width=botWid, height=2, highlightthickness=0).pack(expand=exBool)

    exitB = tkinter.Button(
        button_frame, text="アプリを閉じる", command=exitGiri, width=botWid, height=2, highlightthickness=0).pack(expand=exBool)


# メイン処理 - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
if __name__ == "__main__":

    # 画面処理
    root = tkinter.Tk()
    root.title("定期考査採点アプリ")
    root.geometry("800x400")
    root.configure(bg='white')

    top_activate()

    root.mainloop()
